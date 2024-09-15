from flask import Flask, render_template, request, redirect
import openpyxl
from openpyxl import Workbook
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_PATH = 'attendance.xlsx'

# Ensure Excel exists
def create_excel_if_not_exists():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        # Headers
        ws.append(["Roll No", "Name", "Month", "Total Days", "Days Present", "Attendance %", "Present Dates"])
        wb.save(EXCEL_PATH)

create_excel_if_not_exists()

# Update Excel with attendance data
def update_excel(roll_no, name, month, total_days, present_days, present_dates):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Attendance"]

    # Check if the student and month data already exists
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == roll_no and row[2].value == month:
            row[3].value = total_days
            row[4].value = present_days
            row[5].value = (present_days / total_days) * 100
            row[6].value = ', '.join(present_dates)
            wb.save(EXCEL_PATH)
            return
    
    # If not found, add a new row
    attendance_percentage = (present_days / total_days) * 100
    ws.append([roll_no, name, month, total_days, present_days, attendance_percentage, ', '.join(present_dates)])
    wb.save(EXCEL_PATH)

# Analyze and create a new sheet for students with <50% attendance
def analyze_attendance(month):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Attendance"]
    
    # Create new sheet for the month if needed
    if f"{month}-LowAttendance" not in wb.sheetnames:
        wb.create_sheet(f"{month}-LowAttendance")
    
    low_attendance_sheet = wb[f"{month}-LowAttendance"]
    low_attendance_sheet.append(["Roll No", "Name", "Days Present", "Total Days", "Attendance %", "Present Dates"])
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        roll_no, name, student_month, total_days, present_days, attendance_pct, present_dates = row
        if student_month == month and attendance_pct < 50:
            low_attendance_sheet.append([roll_no, name, present_days, total_days, attendance_pct, present_dates])
    
    wb.save(EXCEL_PATH)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit_attendance():
    roll_no = request.form['roll_no']
    name = request.form['name']
    month = request.form['month']
    total_days = int(request.form['total_days'])
    present_dates = request.form.getlist('present_dates')
    present_days = len(present_dates)
    
    # Update the Excel file
    update_excel(roll_no, name, month, total_days, present_days, present_dates)
    
    # Analyze the attendance
    analyze_attendance(month)
    
    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)
